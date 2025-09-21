from bs4 import BeautifulSoup
import openpyxl
with open("amazon.html",encoding= "utf8") as f:
    html_content = f.read()
    soup = BeautifulSoup(html_content,"html.parser")
    divs = soup.find_all("div",{"class":"s-card-container s-overflow-hidden aok-relative puis-include-content-margin puis s-latency-cf-section s-card-border"})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Product Name"
    ws["B1"] = "Product Price"
    ws["C1"] = "Product Reviews"
    row = 2
    for div in divs:
        product_name_span = div.find("span",{"class": "a-size-medium a-color-base a-text-normal"})
        if product_name_span:
            product_name =product_name_span.text.strip()
        else:
            product_name =""
        product_price_span = div.find("span",{"class": "a-price-whole"})
        if product_price_span:
            product_price = product_price_span.text.strip()
        else:
            product_price =""
            product_reviews_span = div.find("span",{"class": "a-size-base"})
        if product_reviews_span:
            product_reviews = product_reviews_span.text.strip()
        else:
            product_reviews =""
        ws.cell(row=row, coloumn=1, value=product_name)
        ws.cell(row=row, coloumn=2, value=product_price)
        ws.cell(row=row, coloumn=3, value=product_reviews) 
        row += 1
        wb.save("amazon_products.xlsx") 
                      


        


